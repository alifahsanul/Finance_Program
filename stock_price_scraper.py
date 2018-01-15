from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import load_workbook
from datetime import datetime
import pandas
import sys
import pickle
timelist=[]
#timelist.append(datetime.now())
print('this is a program to get IDX stock prices', flush=True)
filename='C:\\Users\\alifahsanul\\Google Drive\\Finance\\Stock and Bond.xlsx'
sheetname="stock"
start_date='5 Nov 2017'
start_date=datetime.strptime(start_date,'%d %b %Y').date()
wb=load_workbook(filename)
ws=wb[sheetname]
#timelist.append(datetime.now())
try:
    date_list=(pandas.read_excel(filename, sheetname=sheetname, header=None)).iloc[0]
except Exception as ex:
    print (ex)
    input('Workbook not found, press enter to exit')
    sys.exit()

#timelist.append(datetime.now())
existing_ticker=(pandas.read_excel(filename, sheetname=sheetname, header=None, skiprows=1))[0]
existing_ticker=[existing_ticker for existing_ticker in existing_ticker if str(existing_ticker)!='nan']
try:
    html=urlopen('https://www.indonesia-investments.com/finance/stocks-bonds/jakarta-composite-index-ihsg/item887')
except Exception as ex:
    print (ex)
    input('Can\'t parse website, check it in your browser')
    sys.exit()

#timelist.append(datetime.now())

bsObj=BeautifulSoup(html,"lxml")
trs=bsObj.find_all(['tr','td'],{'Class':''})

dirtyprice=[]

#timelist.append(datetime.now())

for i in range(len(trs)):
    dirtyprice.append(str(trs[i]))

#timelist.append(datetime.now())

price=[]
for i in range(len(dirtyprice)):
    if '<td class="firstcol">' in dirtyprice[i]:
        price.append((dirtyprice[i+2])[4:-5])

#timelist.append(datetime.now())

price_today=[]
price_yesterday=[]
for i in range (len(price)):
    if i%2==0:
        price_today.append(int(price[i].replace(',','')))
    else:
        price_yesterday.append(int(price[i].replace(',','')))

#timelist.append(datetime.now())

ticker_list=bsObj.find_all("strong",{"Class":""})
dirty_ticker=[]

for i in range (len(ticker_list)):
    dirty_ticker.append(ticker_list[i].get_text())

today=(datetime.strptime((dirtyprice[2])[7:-5],'%d %b %Y')).date()
yesterday=(datetime.strptime((dirtyprice[3])[7:-5],'%d %b %Y')).date()

#timelist.append(datetime.now())
ticker=[]
for i in range(len(dirty_ticker)):
    if len(dirty_ticker[i])==4 and dirty_ticker[i].isupper():
        ticker.append(dirty_ticker[i])
#timelist.append(datetime.now())
if len(ticker)!=len(existing_ticker):
    print('ERROR!!! ticker size is different')
    input('Press enter to exit')
    sys.exit()
#timelist.append(datetime.now())

check_today=0
check_yesterday=0

for j in range(1,len(date_list)):
    try:
        dummy_date=datetime.strptime(str(date_list[j]),'%Y-%m-%d %H:%M:%S').date()
    except Exception as ex:
        print (ex)
        input('Date is wrong')
        sys.exit()
    if dummy_date==yesterday:
        print('getting stock price for', yesterday)
        for i in range(len(ticker)):
            if ticker[i]!=existing_ticker[i]:
                print('ERROR!!! ticker not same, location:',i+1,'th row')
                input('Press enter to exit')
                sys.exit()
            check_yesterday=1
            ws.cell(row=2+i,column=1).value=ticker[i]
            ws.cell(row=2+i, column=j+1).value=int(price_yesterday[i])
    if dummy_date==today:
        print('getting stock price for',today)
        for i in range(len(ticker)):
            if ticker[i]!=existing_ticker[i]:
                print('ERROR!!! ticker not same, location:',i+1,'th row')
                input('Press enter to exit')
                sys.exit()
            check_today=1
            ws.cell(row=2+i,column=1).value=ticker[i]
            ws.cell(row=2+i, column=j+1).value=int(price_today[i])

#timelist.append(datetime.now())

if check_today==0 and check_yesterday==0:
    print('today and yesterday columns not found')
    input('Press enter to exit')
    sys.exit()
elif check_today==0:
    print('today column not found')
    input('Press enter to exit')
    sys.exit()
elif check_yesterday==0:
    print('yesterday column not found')
    input('Press enter to exit')
    sys.exit()


#alltimedata=[None]*365
alltimedata=pickle.load(open('save.p','rb'))
alltimedata[(today-start_date).days]=[ticker,price_today]
alltimedata[(yesterday-start_date).days]=[ticker,price_yesterday]
pickle.dump(alltimedata, open('save.p','wb'))

try:
    wb.save(filename)
except Exception as ex:
    print(ex)
    input('Press enter to exit')
    sys.exit()
print('Parsing stock price succeed')
input('Press enter to exit')