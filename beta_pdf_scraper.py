import numpy
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

filename="C:\\Users\\alifahsanul\\Google Drive\\Finance\\Stock and Bond.xlsx"
sheetname="beta"
pdf_file="C:\\Users\\alifahsanul\\Downloads\\beta.pdf"

wb=load_workbook(filename)
ws=wb[sheetname]

pdfFileObj=open(pdf_file,"rb")
beta=pd.DataFrame(columns=["Raw Beta","Adjusted Beta"])
pdfReader=PyPDF2.PdfFileReader(pdfFileObj)
pageObj=pdfReader.getPage(0)
mytext=pageObj.extractText()
edition=(mytext.split("Edition: ",1)[1])

for i in range(pdfReader.getNumPages()):
    pageObj=pdfReader.getPage(i)
    mytext=pageObj.extractText()
    mytext=mytext.replace("\n","").replace("\r","").replace(" ","")
    textarray=numpy.array(list(mytext))
    for j in range(len(textarray)-14):
        char0=str.isupper(textarray[j])
        char1=str.isupper(textarray[j+1])
        char2=str.isupper(textarray[j+2])
        char3=str.isupper(textarray[j+3])
        char4=str.isupper(textarray[j+4])
        char5=str.isupper(textarray[j+5])
        char6=(textarray[j+6])=="."
        checkchar=not(char0) and char1 and char2 and char3 and char4 and not(char5) and char6
        if checkchar:
            dummy0=("".join([textarray[j+1],textarray[j+2],textarray[j+3],textarray[j+4]]))
            
            dummy1=float("".join([textarray[j+5],textarray[j+6],textarray[j+7],textarray[j+8],textarray[j+9]]))
            
            dummy2=float("".join([textarray[j+10],textarray[j+11],textarray[j+12],textarray[j+13],textarray[j+14]]))
            beta.loc[dummy0]=[dummy2,dummy1]

pdfFileObj.close()

rows=dataframe_to_rows(beta)

for r_idx, row in enumerate(rows,1):
    for c_idx, value in enumerate(row,1):
        ws.cell(row=r_idx, column=c_idx, value=value)

ws.cell(row=1, column=1, value=edition)
wb.save(filename)





