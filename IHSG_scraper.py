import datetime
import pandas as pd
import pandas_datareader.data as web
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# delta_day is the difference between today and start date to take the data
delta_day=3000
start=(datetime.datetime.now()-datetime.timedelta(days=delta_day)).date()
end=datetime.date.today()

df=web.DataReader("^JKSE", "yahoo", start, end)
df=df.drop(["Adj Close"], axis=1)

date=[]
for i in range(len(df)):
    date.append(df.index[i].date())

df["New Date"]=pd.Series(date, index=df.index)
df=df.set_index("New Date")
df=df.reindex_axis(["Volume","Open","Low","High","Close"],axis=1)

rows=dataframe_to_rows(df)

filename="C:\\Users\\alifahsanul\\Google Drive\\Finance\\Stock and Bond.xlsx"
sheetname="IHSG"
wb=load_workbook(filename)
ws=wb[sheetname]

for r_idx, row in enumerate(rows,1):
    for c_idx, value in enumerate(row,1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(filename)
