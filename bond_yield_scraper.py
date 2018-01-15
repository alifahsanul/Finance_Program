from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import load_workbook
from datetime import datetime
import pandas
import sys

print('this is a program to get bond yield', flush=True)
filename='C:\\Users\\alifahsanul\\Google Drive\\Finance\\Stock and Bond.xlsx'
sheetname='bond'
start_date='5 Nov 2017'
start_date=datetime.strptime(start_date,'%d %b %Y').date()
wb=load_workbook(filename)
ws=wb[sheetname]
try:
    date_list=(pandas.read_excel(filename, sheetname=sheetname, header=None)).iloc[0]
except Exception as ex:
    print (ex)
    input('Workbook not found, press enter to exit')
    sys.exit()

existing_maturity=(pandas.read_excel(filename, sheetname=sheetname, header=None, skiprows=1))[0]

existing_maturity=[existing_maturity for existing_maturity in existing_maturity if str(existing_maturity)!='nan']

print('parsing', flush=True)
try:
    html=urlopen('https://asianbondsonline.adb.org/indonesia.php')
except Exception as ex:
    print (ex)
    input('Can\'t parse website, check it in your browser')
    sys.exit()
bsObj=BeautifulSoup(html,"lxml")
trs=bsObj.find_all(['tr','td'],{'Class':''})

dirty_bond=[]

for i in range(len(trs)):
    dirty_bond.append(str(trs[i]))

bond_yield=[]
maturity=[]
for i in range(len(dirty_bond)):
    if '<td class="name">' in dirty_bond[i]:
        if '<td class="latest_close">' in dirty_bond[i+2]:
            bond_yield.append(float(((dirty_bond[i+2])[25:-6]).replace(',','')))
        if '<tr>' not in dirty_bond[i]:
            maturity_criteria1='Year' in dirty_bond[i]
            maturity_criteria2='IDR' in dirty_bond[i]
            maturity_criteria3='JIBOR' in dirty_bond[i]
            maturity_criteria=maturity_criteria1 or maturity_criteria2 or maturity_criteria3
            if maturity_criteria:
                maturity.append((dirty_bond[i])[17:-5])

today=bsObj.find_all('span',{'class':'datetime'})
for tag in today:
    today=tag.text.strip()
today=(datetime.strptime(today, '%B %d, %Y')).date()
print('getting bond yield for',today)

if len(maturity)!=len(existing_maturity):
    print('ERROR!!! maturity size is different')
    input('Press enter to exit')
    sys.exit()

check_maturity=0

for j in range(1,len(date_list)):
    try:
        dummy_date=datetime.strptime(str(date_list[j]),'%Y-%m-%d %H:%M:%S').date()
    except Exception as ex:
        print (ex)
        input('Date is wrong')
        sys.exit()
    if dummy_date==today:
        check_maturity=1
        for i in range(len(maturity)):
            if maturity[i]!=existing_maturity[i]:
                print('ERROR!!! maturity is not same, location:',i+1,'th row')
                print(existing_maturity[i])
                input('Press enter to exit')
                sys.exit()
            ws.cell(row=2+i,column=1).value=maturity[i]
            ws.cell(row=2+i, column=j+1).value=bond_yield[i]

if check_maturity!=1:
    print('Maturity date column not found')
    input('Press enter to exit')
    sys.exit()


try:
    wb.save(filename)
except Exception as ex:
    print(ex)
    input('Press enter to exit')
    sys.exit()


print('Parsing bond yield succeed')
input('Press enter to exit')




